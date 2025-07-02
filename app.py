from flask import Flask, render_template, request, redirect, flash, send_file
import os
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import io

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "segredo")  # melhor usar variável de ambiente

# Função para criar arquivo Excel caso não exista
def criar_arquivo_excel(caminho):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append([
        "Nome/Razão Social", "Tipo Documento", "Documento", "Telefone", "E-mail",
        "CEP", "Rua", "Número", "Bairro", "Cidade", "Bem", "Valor", "Parcela",
        "Status", "Origem"
    ])
    wb.save(caminho)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        try:
            nome_razao = request.form.get("nome_razao", "").strip()
            tipo_doc = request.form.get("tipo_doc", "").strip()
            documento = request.form.get("documento", "").strip()
            telefone = request.form.get("telefone", "").strip()
            email = request.form.get("email", "").strip()
            cep = request.form.get("cep", "").strip()
            rua = request.form.get("rua", "").strip()
            numero = request.form.get("numero", "").strip()
            bairro = request.form.get("bairro", "").strip()
            cidade = request.form.get("cidade", "").strip()
            bem = request.form.get("bem", "").strip()
            valor_str = request.form.get("valor", "0").replace(",", ".").strip()
            valor = float(valor_str) if valor_str else 0.0
            status = request.form.get("status", "").strip()
            origem = request.form.get("origem", "").strip()

            # Validações básicas
            if not (nome_razao and tipo_doc and documento and telefone and cep and bem and valor > 0):
                flash("Preencha todos os campos obrigatórios corretamente.", "danger")
                return redirect("/")

            # Cálculo da parcela
            if bem == "Automóvel":
                if valor < 65000:
                    flash("Valor mínimo para Automóvel é R$ 65.000,00", "danger")
                    return redirect("/")
                parcela = round((valor * 1.16) / 84, 2)
            elif bem == "Imóvel":
                if valor < 200000:
                    flash("Valor mínimo para Imóvel é R$ 200.000,00", "danger")
                    return redirect("/")
                parcela = round((valor * 1.22) / 220, 2)
            else:
                parcela = 0

            arquivo = "leads.xlsx"
            if not os.path.exists(arquivo):
                criar_arquivo_excel(arquivo)

            wb = openpyxl.load_workbook(arquivo)
            ws = wb["Leads"]
            ws.append([
                nome_razao, tipo_doc, documento, telefone, email, cep, rua,
                numero, bairro, cidade, bem, valor, parcela, status, origem
            ])
            wb.save(arquivo)

            flash(f"Lead cadastrado com sucesso! Parcela: R$ {parcela:.2f}", "success")
            return redirect("/")
        except Exception as e:
            flash(f"Erro ao processar os dados: {e}", "danger")
            return redirect("/")
    return render_template("index.html")


@app.route("/leads")
def visualizar_leads():
    leads = []
    arquivo = "leads.xlsx"

    filtros = {
        "nome": request.args.get("nome", "").lower(),
        "telefone": request.args.get("telefone", "").lower(),
        "email": request.args.get("email", "").lower(),
        "cidade": request.args.get("cidade", "").lower(),
        "status": request.args.get("status", "").lower()
    }

    if os.path.exists(arquivo):
        wb = openpyxl.load_workbook(arquivo)
        if "Leads" in wb.sheetnames:
            ws = wb["Leads"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                nome = str(row[0] or "").lower()
                telefone = str(row[3] or "").lower()
                email = str(row[4] or "").lower()
                cidade = str(row[9] or "").lower()
                status = str(row[13] or "").lower()

                if (
                    (not filtros["nome"] or filtros["nome"] in nome) and
                    (not filtros["telefone"] or filtros["telefone"] in telefone) and
                    (not filtros["email"] or filtros["email"] in email) and
                    (not filtros["cidade"] or filtros["cidade"] in cidade) and
                    (not filtros["status"] or filtros["status"] == status)
                ):
                    leads.append(row)

    return render_template("leads.html", leads=leads)


@app.route("/exportar-pdf")
def exportar_pdf():
    arquivo = "leads.xlsx"
    buffer = io.BytesIO()

    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    c.setFont("Helvetica-Bold", 14)
    y = height - 40

    if os.path.exists(arquivo):
        wb = openpyxl.load_workbook(arquivo)
        ws = wb["Leads"]

        c.drawString(40, y, "Leads Exportados - Raros Capital")
        y -= 30
        c.setFont("Helvetica", 10)

        headers = ["Nome", "Telefone", "E-mail", "Cidade", "Status", "Origem", "Valor", "Parcela"]
        c.drawString(40, y, " | ".join(headers))
        y -= 20

        for row in ws.iter_rows(min_row=2, values_only=True):
            if y < 50:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica", 10)
            # Seleciona os campos principais para exibir no PDF
            dados_pdf = [
                str(row[0] or ""),    # Nome/Razão Social
                str(row[3] or ""),    # Telefone
                str(row[4] or ""),    # E-mail
                str(row[9] or ""),    # Cidade
                str(row[13] or ""),   # Status
                str(row[14] or ""),   # Origem
                f"R$ {row[11]:,.2f}" if row[11] else "R$ 0,00",  # Valor formatado
                f"R$ {row[12]:,.2f}" if row[12] else "R$ 0,00",  # Parcela formatada
            ]
            linha = " | ".join(dados_pdf)
            c.drawString(40, y, linha)
            y -= 15

        c.save()
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name="leads.pdf", mimetype="application/pdf")
    else:
        flash("Arquivo de leads não encontrado.", "danger")
        return redirect("/leads")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
